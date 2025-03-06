# Это не совсем старый конвертер. Тут посто сокращённые названия единиц, которые отличны от названий в программе
import math
import os

from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side
from AllUnitsDB import AllUnits, main

# Density
def density_converter(value, from_unit, to_unit, units):
    """Конвертер плотности."""
    # units = AllUnits.density_units
    # Проверка, поддерживаются ли единицы измерения
    if from_unit not in units or to_unit not in units:
        raise ValueError("Неизвестная единица измерения плотности.")

    # Конвертация
    value_in_kgm3 = value * units[from_unit]
    return value_in_kgm3 / units[to_unit]

# Distance
def distance_converter(value, from_unit, to_unit, units):
    """Конвертер расстояния."""
    print("Зашли в функцию дистансе")
    # units = AllUnits.distance_units
    print(f"UNITS{units}")
    value_in_meters = value * units[from_unit]
    return value_in_meters / units[to_unit]



# Energy
def energy_converter(value, from_unit, to_unit, units):
    """Конвертер энергии."""
    # units = AllUnits.energy_units
    value_in_joules = value * units[from_unit]
    return value_in_joules / units[to_unit]


# Flow (Volume Flow Rate)
def flow_converter(value, from_unit, to_unit, units):
    """Конвертер объемного расхода."""
    # units = AllUnits.flow_units
    value_in_m3s = value * units[from_unit]
    return value_in_m3s / units[to_unit]

# Force
def force_converter(value, from_unit, to_unit, units):
    """Конвертер силы."""
    # units = AllUnits.force_units
    value_in_newtons = value * units[from_unit]
    return value_in_newtons / units[to_unit]


# Light (Illuminance)
def illuminance_converter(value, from_unit, to_unit, units):
    """Конвертер освещенности."""
    # units = AllUnits.light_units
    value_in_lux = value * units[from_unit]
    return value_in_lux / units[to_unit]


# Mass
def mass_converter(value, from_unit, to_unit, units):
    """Конвертер массы."""
    # units = AllUnits.mass_units
    value_in_kg = value * units[from_unit]
    return value_in_kg / units[to_unit]

# Power
def power_converter(value, from_unit, to_unit, units):
    """Конвертер мощности."""
    # units = AllUnits.power_units
    value_in_watts = value * units[from_unit]
    return value_in_watts / units[to_unit]

# Pressure
def pressure_converter(value, from_unit, to_unit, units):
    """Конвертер давления (расширенный)."""
    # units = AllUnits.pressure_units
    value_in_pascals = value * units[from_unit]
    return value_in_pascals / units[to_unit]

def speed_converter(value, from_unit, to_unit, units):
    """Конвертер скорости (расширенный)."""
    # units = AllUnits.speed_units
    value_in_ms = value * units[from_unit]
    return value_in_ms / units[to_unit]

# Temperature
def temperature_converter(value, from_unit, to_unit, units):
    """Конвертер температуры."""
    if from_unit not in units.keys() or to_unit not in units.keys():
        raise ValueError("Неизвестная единица измерения температуры.")
    celsius = units[from_unit][0](value)
    return units[to_unit][1](celsius)


# Time
def time_converter(value, from_unit, to_unit, units):
    """Конвертер времени."""
    # units = AllUnits.time_units
    # Конвертация
    value_in_seconds = value * units[from_unit]
    return value_in_seconds / units[to_unit]

# Torque
def torque_converter(value, from_unit, to_unit, units):
    """Конвертер крутящего момента."""
    # units = AllUnits.torque_units
    value_in_Nm = value * units[from_unit]
    return value_in_Nm / units[to_unit]

# Volume
def volume_converter(value, from_unit, to_unit, units):
    """Конвертер объема."""
    # units = AllUnits.volume_units
    value_in_m3 = value * units[from_unit]
    return value_in_m3 / units[to_unit]

# Volume - Dry (US)
def dry_volume_converter(value, from_unit, to_unit, units):
    """Конвертер сухого объема (US)."""
    # units = AllUnits.dry_volume_units

    value_in_bushels = value * units[from_unit]
    return value_in_bushels / units[to_unit]

# Acceleration
def acceleration_converter(value, from_unit, to_unit, units):
    """Конвертер ускорения."""
    # units = AllUnits.acceleration_units
    value_in_ms2 = value * units[from_unit]
    return value_in_ms2 / units[to_unit]

# Amt. of Substance (Amount of Substance)
def amount_converter(value, from_unit, to_unit, units):
    """Конвертер количества вещества."""
    # units = AllUnits.amount_unit
    value_in_mol = value * units[from_unit]
    return value_in_mol / units[to_unit]

# Angle
def angle_converter(value, from_unit, to_unit, units):
    """Конвертер углов."""
    # units = AllUnits.angle_unit
    value_in_radians = value * units[from_unit]
    return value_in_radians / units[to_unit]

# Area
def area_converter(value, from_unit, to_unit, units):
    """Конвертер площади."""
    # units = AllUnits.area_unit
    value_in_m2 = value * units[from_unit]
    return value_in_m2 / units[to_unit]

# Computer (Data Storage)
def data_converter(value, from_unit, to_unit, units):
    """Конвертер объема данных."""
    # units = AllUnits.computer_units
    value_in_bytes = value * units[from_unit]
    return value_in_bytes / units[to_unit]

# Concentration (Mass Concentration)
def concentration_converter(value, from_unit, to_unit, units):
  """Конвертер концентрации (молярной)."""
  # units = AllUnits.concentration_units
  value_in_kmolm3 = value * units[from_unit]
  return value_in_kmolm3 / units[to_unit]

# Custom
def custom_converter(value, from_unit, to_unit, custom_units):
    """
    Конвертирует значение из одной пользовательской единицы в другую.

    :param value: Значение для конвертации.
    :param from_unit: Исходная единица измерения.
    :param to_unit: Целевая единица измерения.
    :param custom_units: Словарь с пользовательскими единицами.
    :return: Результат конвертации.
    """
    try:
        # Получаем данные для исходной единицы
        from_unit_data = custom_units.get(from_unit)
        if not from_unit_data:
            raise ValueError(f'Исходная единица "{from_unit}" не найдена.')

        # Получаем данные для целевой единицы
        to_unit_data = custom_units.get(to_unit)
        if not to_unit_data:
            raise ValueError(f'Целевая единица "{to_unit}" не найдена.')

        # Проверяем, что обе единицы используют одну и ту же базовую единицу
        if from_unit_data['base_unit'] != to_unit_data['base_unit']:
            raise ValueError('Единицы измерения должны использовать одну и ту же базовую единицу.')

        # Конвертируем значение в базовую единицу
        value_in_base = value * from_unit_data['value']

        # Конвертируем значение из базовой единицы в целевую
        result = value_in_base / to_unit_data['value']

        return result

    except Exception as e:
        raise ValueError(f'Ошибка при конвертации: {str(e)}')

def test_converter(units, type, function):
    """Тестирует pressure_converter, преобразуя во все единицы измерения."""

    value = float(input("Введите значение: "))
    from_unit = input("Введите исходную единицу измерения: ")

    print(f"ALL UNITS: {units}")

    if from_unit not in units:  # Проверяем, есть ли такая единица
        print("Ошибка: Недопустимая исходная единица измерения.")
        return

    # 1. Сортировка ключей в список:

    sorted_units_keys = sorted(units.keys(), key=lambda s: s.lower())
    print("Отсортированные ключи:", sorted_units_keys)

    # 2. Сортировка и итерация по отсортированным ключам:

    print("\nИтерация по отсортированным ключам:")
    for key in sorted_units_keys:
        print(f"{key}: {units[key]}")

    # 3. Получение отсортированного словаря (если требуется):
    #    (Обратите внимание, что это создаст *новый* отсортированный словарь.
    #     Исходный словарь units не изменится.)

    import collections

    sorted_units = collections.OrderedDict((key, units[key]) for key in sorted_units_keys)
    excel_creater(value, from_unit, sorted_units, type, function)




def excel_creater(value, from_unit, to_units, type, conversion_function):
    from openpyxl import Workbook

    """
    Создаёт Excel-файл с результатами конвертации.

    :param value: Значение для конвертации.
    :param from_unit: Начальная единица измерения.
    :param to_units: Список единиц измерения, в которые нужно конвертировать.
    """
    # Создаём новую книгу Excel и выбираем активный лист
    wb = Workbook()
    ws = wb.active

    # Заголовки таблицы
    ws.append(["Значение", "Начальная величина", "Значение", "Сконвертированная величина"])

    # Конвертируем значение в каждую из указанных единиц измерения
    for to_unit in to_units:
        try:
            print(f"value, from_unit, to_unit: {value, from_unit, to_unit}")
            converted_value = conversion_function(value, from_unit, to_unit)
            ws.append([value, from_unit, converted_value, to_unit])
            print(f"{value} {from_unit} = {converted_value} {to_unit}")
        except Exception as e:
            print(f"Ошибка при конвертации {from_unit} в {to_unit}: {e}")

    # Настраиваем стили для красивого вида
    header_style = NamedStyle(name='header')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    border_style = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    header_style.border = border_style

    cell_style = NamedStyle(name='cell')
    cell_style.alignment = Alignment(horizontal='left', vertical='center')
    cell_style.border = border_style

    for cell in ws[1]:  # Применяем стиль к заголовкам
        cell.style = header_style

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.style = cell_style

    # Автоматическое изменение ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    filename = f"conversion_results_{type}.xlsx"
    # Проверка существования папки
    if not os.path.exists("Conversions"):
        # Создание папки, если она не существует
        os.makedirs("Conversions")
        print(f"Папка Conversions создана.")
    else:
        print(f"Папка Conversions уже существует.")
    wb.save(f"Conversions\\{filename}")
    print(f"Файл '{filename}' успешно создан.")


# # Пример использования
# value = 1  # Значение для конвертации
# from_unit = "atm"  # Начальная единица измерения
# to_units = ["Pa", "kPa", "MPa", "bar", "mmHg", "PSI"]  # Список единиц для конвертации
#
# excel_creater(value, from_unit, to_units)

if __name__ == "__main__":
    AllUnits = main()
    unit = AllUnits.pressure_units
    test_converter(unit, "pressure1", pressure_converter)